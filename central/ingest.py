"""
central.ingest
==============
Generic, idempotent importer: every dataset in the manifest becomes a table.

Design goals the request named:
  * "all existing content in a central DB"  -> ingests every content sheet.
  * "really easy to add content"            -> re-run to upsert edited sheets
                                               (stable row ids, no duplicates),
                                               plus `add` / `template` CLI verbs.
  * "let other devs api-ize off the db"     -> writes a self-describing registry
                                               (_datasets, _columns) the API reads,
                                               and per-column visibility flags.

Tables get four housekeeping columns: _row_id (PK), _source_file, _source_sheet,
_imported_at. Everything else comes from the sheet headers (slugged), with the
original header preserved in the _columns registry so the API can show devs the
human label.

CLI:
    python -m central.ingest build   --uploads /path/to/xlsx_dir
    python -m central.ingest list
    python -m central.ingest template --table studios
    python -m central.ingest add     --table studios name="New Studio" city="Crozet" country="USA"
"""

from __future__ import annotations

import argparse
import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from . import manifest as M
from .dbconn import connect, local_path

# --- registry (self-describing metadata the API serves) --------------------
REGISTRY_DDL = """
CREATE TABLE IF NOT EXISTS _datasets (
    tbl         TEXT PRIMARY KEY,
    domain      TEXT,
    source_file TEXT,
    source_sheet TEXT,
    visibility  TEXT,
    row_count   INTEGER,
    description TEXT,
    updated_at  TEXT
);
CREATE TABLE IF NOT EXISTS _columns (
    tbl       TEXT,
    column    TEXT,
    label     TEXT,        -- original sheet header
    ordinal   INTEGER,
    is_public INTEGER,
    PRIMARY KEY (tbl, column)
);
"""


def _blank(v) -> bool:
    return v is None or (isinstance(v, str) and not v.strip())


def _row_id(table: str, key_slugs: list[str], record: dict) -> str:
    parts = [str(record.get(k, "")) for k in key_slugs] if key_slugs else []
    if not any(p.strip() for p in parts):          # no usable key -> hash whole row
        parts = [f"{k}={record[k]}" for k in sorted(record)]
    blob = (table + "|" + "|".join(parts)).encode("utf-8")
    return hashlib.sha1(blob).hexdigest()[:16]


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def ingest_dataset(conn, ds: M.Dataset, uploads: Path) -> int:
    path = uploads / ds.source_file
    if not path.exists():
        print(f"  ! missing file, skipped: {ds.source_file}")
        return 0
    wb = load_workbook(path, read_only=True, data_only=True)
    if ds.sheet not in wb.sheetnames:
        print(f"  ! missing sheet '{ds.sheet}' in {ds.source_file}")
        return 0
    ws = wb[ds.sheet]

    it = ws.iter_rows(values_only=True)
    try:
        raw_header = next(it)
    except StopIteration:
        return 0

    used: set[str] = set()
    slugs, labels = [], []
    for h in raw_header:
        label = str(h).strip() if h is not None else ""
        if label == "":
            slugs.append(None); labels.append(None); continue
        s = M.slugify(label, used)
        slugs.append(s); labels.append(label)

    key_slugs = [M.slugify(k) for k in ds.key_columns]  # deterministic (no `used`)

    # (re)create table with housekeeping + data columns
    real = {s for s in slugs if s in M.REAL_COLUMNS}
    col_defs = ['"_row_id" TEXT PRIMARY KEY', '"_source_file" TEXT',
                '"_source_sheet" TEXT', '"_imported_at" TEXT']
    for s in slugs:
        if s is None:
            continue
        col_defs.append(f'"{s}" {"REAL" if s in real else "TEXT"}')
    conn.execute(f'DROP TABLE IF EXISTS "{ds.table}"')
    conn.execute(f'CREATE TABLE "{ds.table}" ({", ".join(col_defs)})')

    data_slugs = [s for s in slugs if s is not None]
    now = datetime.now(timezone.utc).isoformat()
    n = 0
    for raw in it:
        if not any(not _blank(v) for v in raw):
            continue
        record = {}
        for s, val in zip(slugs, raw):
            if s is None:
                continue
            if s in real:
                record[s] = _num(val)
            else:
                record[s] = None if val is None else str(val).strip()
        rid = _row_id(ds.table, key_slugs, record)
        cols = ["_row_id", "_source_file", "_source_sheet", "_imported_at", *data_slugs]
        vals = [rid, ds.source_file, ds.sheet, now, *[record.get(s) for s in data_slugs]]
        collist = ", ".join('"' + c + '"' for c in cols)
        ph = ", ".join("?" for _ in cols)
        conn.execute(
            f'INSERT OR REPLACE INTO "{ds.table}" ({collist}) VALUES ({ph})',
            vals,
        )
        n += 1

    # registry
    conn.execute(
        """INSERT OR REPLACE INTO _datasets
           (tbl, domain, source_file, source_sheet, visibility, row_count, description, updated_at)
           VALUES (?,?,?,?,?,?,?,?)""",
        (ds.table, ds.domain, ds.source_file, ds.sheet, ds.visibility, n, ds.description, now),
    )
    conn.execute('DELETE FROM _columns WHERE tbl=?', (ds.table,))
    ordinal = 0
    for s, label in zip(slugs, labels):
        if s is None:
            continue
        conn.execute(
            "INSERT OR REPLACE INTO _columns (tbl, column, label, ordinal, is_public) VALUES (?,?,?,?,?)",
            (ds.table, s, label, ordinal, 0 if M.is_private_column(s) else 1),
        )
        ordinal += 1
    conn.commit()
    print(f"  · {ds.table:28s} {n:5d} rows  [{ds.visibility}]  <- {ds.source_file} :: {ds.sheet}")
    return n


def build(uploads: Path) -> None:
    conn = connect()
    conn.executescript(REGISTRY_DDL) if hasattr(conn, "executescript") else [
        conn.execute(s) for s in REGISTRY_DDL.split(";") if s.strip()
    ]
    conn.commit()
    total = 0
    print(f"Building central DB at {local_path()}")
    for ds in M.DATASETS:
        total += ingest_dataset(conn, ds, uploads)
    # checkpoint so the file is ready for `turso db create --from-file`
    try:
        conn.execute("PRAGMA wal_checkpoint(TRUNCATE);")
    except Exception:
        pass
    conn.commit()
    n_tables = conn.execute("SELECT COUNT(*) FROM _datasets").fetchone()[0]
    print(f"\nDone: {n_tables} datasets, {total} rows total.")
    _print_skips()


def _print_skips() -> None:
    print("\nDeliberately skipped (spreadsheet machinery / duplicates):")
    for (f, sh), why in M.SKIP.items():
        print(f"  – {f} :: {sh}  ({why})")


def list_datasets() -> None:
    conn = connect()
    rows = conn.execute(
        "SELECT domain, tbl, visibility, row_count, description FROM _datasets ORDER BY domain, tbl"
    ).fetchall()
    dom = None
    for r in rows:
        if r["domain"] != dom:
            dom = r["domain"]; print(f"\n[{dom}]")
        vis = "" if r["visibility"] == "public" else "  (restricted)"
        print(f"  {r['tbl']:28s} {r['row_count']:5d}  {r['description']}{vis}")


def template(table: str) -> None:
    conn = connect()
    cols = conn.execute(
        "SELECT column, label, is_public FROM _columns WHERE tbl=? ORDER BY ordinal", (table,)
    ).fetchall()
    if not cols:
        print(f"No such table: {table}"); return
    print(f"# add a row to '{table}' — column=value pairs:\n")
    print("python -m central.ingest add --table " + table + " \\")
    for c in cols:
        tag = "" if c["is_public"] else "   # private"
        print(f'    {c["column"]}="" \\{tag}')


def add_row(table: str, values: dict[str, str]) -> None:
    conn = connect()
    known = {r["column"] for r in conn.execute("SELECT column FROM _columns WHERE tbl=?", (table,))}
    if not known:
        print(f"No such table: {table}"); return
    unknown = set(values) - known
    if unknown:
        print(f"Unknown column(s) for {table}: {', '.join(sorted(unknown))}")
        print("Run:  python -m central.ingest template --table " + table); return
    now = datetime.now(timezone.utc).isoformat()
    rid = hashlib.sha1((table + "|manual|" + json.dumps(values, sort_keys=True) + now).encode()).hexdigest()[:16]
    cols = ["_row_id", "_source_file", "_source_sheet", "_imported_at", *values.keys()]
    vals = [rid, "manual", "manual", now, *values.values()]
    collist = ", ".join('"' + c + '"' for c in cols)
    ph = ", ".join("?" for _ in cols)
    conn.execute(f'INSERT INTO "{table}" ({collist}) VALUES ({ph})', vals)
    conn.execute("UPDATE _datasets SET row_count = row_count + 1 WHERE tbl=?", (table,))
    conn.commit()
    print(f"Added row {rid} to {table}.")


# --- CLI -------------------------------------------------------------------
def main(argv=None) -> None:
    ap = argparse.ArgumentParser(prog="central.ingest")
    sub = ap.add_subparsers(dest="cmd", required=True)

    b = sub.add_parser("build", help="ingest all datasets from a folder of xlsx files")
    b.add_parser = None
    b.add_argument("--uploads", required=True)

    sub.add_parser("list", help="list datasets and row counts")

    t = sub.add_parser("template", help="print an add-row command for a table")
    t.add_argument("--table", required=True)

    a = sub.add_parser("add", help="add one row: --table T col=value col=value ...")
    a.add_argument("--table", required=True)
    a.add_argument("pairs", nargs="*")

    args = ap.parse_args(argv)
    if args.cmd == "build":
        build(Path(args.uploads).expanduser())
    elif args.cmd == "list":
        list_datasets()
    elif args.cmd == "template":
        template(args.table)
    elif args.cmd == "add":
        values = {}
        for p in args.pairs:
            if "=" not in p:
                print(f"Ignoring '{p}' (need col=value)"); continue
            k, v = p.split("=", 1)
            values[k.strip()] = v
        add_row(args.table, values)


if __name__ == "__main__":
    main()
